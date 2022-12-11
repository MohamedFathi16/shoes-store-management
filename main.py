from tkinter import*
import datetime
from tkinter.tix import IMAGETEXT
from tkinter import ttk
from PIL import Image,ImageTk
import openpyxl 
from openpyxl import Workbook



pro = Tk() # عبارة عن CLASS داخل مكتبة TKINTER 

now= datetime.datetime.now()
date= now.strftime("%Y-%m-%d") 
pro.geometry('900x522') # مساحة نافذة العرض عند تنفيذ البرنامج 
pro.iconbitmap('images/shoes1670416194.ico') # وضع ايقونة للبرنامج 
pro.title('shoes store') # اسم المشروع 
pro.config(background='#fce294')
#====================== data exel                      ==============
wb = Workbook()
ws=wb.active
ws.title='sells'
ws["A1"]='اسم العميل'
ws["B1"]= 'رقم الهاتف'
ws["C1"]='عنوان العميل '
ws["D1"]='الحساب الكلي'
ws["E1"]= 'التاريخ'
wb.save('mohamed.xlsx')

def save():
    name =enname.get()
    phone=enphone.get()
    total2=entotal.get()
    addresss=enaddres.get()
    datte=endatel.get()

    excel = openpyxl.load_workbook('mohamed.xlsx')
    file = excel.active
    file.cell(column=1,row=file.max_row+1,value=name)
    file.cell(column=2,row=file.max_row,value=phone)
    file.cell(column=3,row=file.max_row,value=total2)
    file.cell(column=3,row=file.max_row,value=addresss)
    file.cell(column=2,row=file.max_row,value=datte)
    excel.save('mohamed.xlsx')

#====================== function of buying ==========
def bil ():
    global enname 
    global enaddres 
    global endatel
    global enphone
    global entotal
    






    pro.geometry('1250x522')
    F4 =Frame(pro, bg='#e9cd5e',width=400 , height=715,bd=1,relief=GROOVE)
    F4.place(x=1120,y=1)

    name1 = Label(F4,text=' : اسم المشتري', bg='#92844d', fg='white', font='18')
    name1.place(x=270,y=50)

    enname = Entry(F4, width=25,font=('Tajwal 14'),justify=CENTER)
    enname.place(x=55,y=100)

    address1= Label(F4, text=' : عنوان المشتري ' ,bg='#92844d', fg='white', font='18' )
    address1.place(x=250, y=150)

    enaddres = Entry(F4, width=25,font=('Tajwal 14'),justify=CENTER)
    enaddres.place(x=55,y=200)

    phone1= Label(F4, text='   : هاتف المشتري ' ,bg='#92844d', fg='white', font='18' )
    phone1.place(x=250, y=240)

    enphone = Entry(F4, width=25,font=('Tajwal 14'),justify=CENTER)
    enphone.place(x=55,y=300)

    total1= Label(F4, text='   :الحساب الكلي ' ,bg='#92844d', fg='white', font='18' )
    total1.place(x=255, y=350)

    entotal =Entry(F4, width=25,font=('Tajwal 14'),justify=CENTER)
    entotal.place(x=55,y=400)

    data1= Label(F4, text='   :   التاريخ' ,bg='#92844d', fg='white', font='18' )
    data1.place(x=290, y=450)

    endatel =Entry(F4, width=25,font=('Tajwal 14'),justify=CENTER)
    endatel.place(x=55,y=500)

    add_button = Button(F4, text='حفظ الفاتورة', width=31, cursor='hand2',bg='#EDDBC0',command=save)
    add_button.place(x=80,y=540)

    add_button = Button(F4, text='افراغ الحقول ', width=31, cursor='hand2',bg='#EDDBC0',command=clea1)
    add_button.place(x=80,y=570)

    add_button = Button(F4, text='بحث عن مشتري', width=31, cursor='hand2',bg='#EDDBC0')
    add_button.place(x=80,y=600)

    add_button = Button(F4, text='حذف الفاتورة', width=31, cursor='hand2',bg='#EDDBC0')
    add_button.place(x=80,y=630)







    #===============          =============

    total = 0
    for item in trv.get_children():
        trv.delete(item)
    for i in range(len(sb)):
        if(int(sb[i].get())>0):
            price = int(sb[i].get())*menu[i][1]
            total= total + price
            myst=(str(menu[i][1]),str(sb[i].get()),str(price))
            trv.insert("",'end',iid=i,text=menu[i][0],values=myst)
    final =total
    entotal.insert('1', str(final) + '$')  
    endatel.insert('1', str(date))
def clea():
    for item in trv.get_children():
        trv.delete(item)
    enname.delete('0',END)
    enphone.delete('0',END)
    entotal.delete('0',END)
    endatel.delete('0',END)
    enaddres.delete('0',END)

def clea1():
    enname.delete('0',END)
    enphone.delete('0',END)
    entotal.delete('0',END)
    endatel.delete('0',END)
    enaddres.delete('0',END)








# FRAME  : تقسيم الواجهة
# هنستدعي فرام من مكتبة tkinter اسمه frame CALSS
#  ========= FRAME 1 =========== 
F1 = Frame(pro , bg='#EE9A4D', width=600, height=715)
F1.place(x=1,y=1)  # X==> يبتعد من ليسار بمقدار ||| Y==> يبتعد من الاعلى بمقدار|||| PLACE ==> تحديد مكان الفيرام فى الواجهة

#  ============= IAMGES =========

img=Image.open('images/5.png')
img=img.resize((85,85))
img=ImageTk.PhotoImage(img)

im2g=Image.open('images/4.png')
im2g=im2g.resize((85,85))
im2g=ImageTk.PhotoImage(im2g)

im3g=Image.open('images/8.png')
im3g=im3g.resize((85,85))
im3g=ImageTk.PhotoImage(im3g)

im4g=Image.open('images/99.png')
im4g=im4g.resize((85,85))
im4g=ImageTk.PhotoImage(im4g)

im5g=Image.open('images/66.png')
im5g=im5g.resize((85,85))
im5g=ImageTk.PhotoImage(im5g)

im6g=Image.open('images/555.png')
im6g=im6g.resize((85,85))
im6g=ImageTk.PhotoImage(im6g)

im7g=Image.open('images/44.png')
im7g=im7g.resize((85,85))
im7g=ImageTk.PhotoImage(im7g)

im8g=Image.open('images/484.png')
im8g=im8g.resize((85,85))
im8g=ImageTk.PhotoImage(im8g)

im9g=Image.open('images/145.png')
im9g=im9g.resize((85,85))
im9g=ImageTk.PhotoImage(im9g)

im10g=Image.open('images/3302.png')
im10g=im10g.resize((85,85))
im10g=ImageTk.PhotoImage(im10g)

im11g=Image.open('images/25544.png')
im11g=im11g.resize((85,85))
im11g=ImageTk.PhotoImage(im11g)

im12g=Image.open('images/1010.png')
im12g=im12g.resize((85,85))
im12g=ImageTk.PhotoImage(im12g)

im13g=Image.open('images/120.png')
im13g=im13g.resize((85,85))
im13g=ImageTk.PhotoImage(im13g)

im14g=Image.open('images/753.png')
im14g=im14g.resize((85,85))
im14g=ImageTk.PhotoImage(im14g)

im15g=Image.open('images/3333.png')
im15g=im15g.resize((85,85))
im15g=ImageTk.PhotoImage(im15g)

im16g=Image.open('images/7777.png')
im16g=im16g.resize((85,85))
im16g=ImageTk.PhotoImage(im16g)

im17g=Image.open('images/987.png')
im17g=im17g.resize((85,85))
im17g=ImageTk.PhotoImage(im17g)





#==========================              =========================================


title = Label(F1 , text='shoes store :', font=('arial 16 bold'), fg='red', bg='#FFA500') # عنوان أوأسم المحل
title.place(x=200 ,y=0) # مكان العنوان فى الفيرام 

menu1 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=img, text='ADIDAS', compound=TOP)
menu1.place(x=30 , y=45)

menu2 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im2g, text='NIKE', compound=TOP) # make a button 
menu2.place(x=160 , y=45)  # picture place 

menu3 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im3g, text='PUMA', compound=TOP)
menu3.place(x=300 , y=45)

menu4 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im4g, text='VL', compound=TOP)
menu4.place(x=450 , y=45)

menu5 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im5g, text='VL', compound=TOP)
menu5.place(x=30 , y=220)

menu6 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im6g, text='VL', compound=TOP)
menu6.place(x=160 , y=220)

menu7 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im7g, text='VL', compound=TOP)
menu7.place(x=300 , y=220)

menu8 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im8g, text='BOOT', compound=TOP)
menu8.place(x=450 , y=220)

menu9 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im9g, text='VL', compound=TOP)
menu9.place(x=30 , y=380)

menu10 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im10g, text='VL', compound=TOP)
menu10.place(x=160 , y=380)

menu11 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im11g, text='VL', compound=TOP)
menu11.place(x=300 , y=380)

menu12 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im12g, text='VL', compound=TOP)
menu12.place(x=450 , y=380)

menu13 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im13g, text='VL', compound=TOP)
menu13.place(x=450 , y=380)

menu14 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im14g, text='VL', compound=TOP)
menu14.place(x=30 , y=535)

menu15 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im15g, text='VL', compound=TOP)
menu15.place(x=160 , y=535)

menu16 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im16g, text='VL', compound=TOP)
menu16.place(x=300 , y=535)

menu17 = Button(F1, width=90 , bg='#FFA62F',bd=1,relief=SOLID,cursor='hand2', height=100,image=im17g, text='VL', compound=TOP)
menu17.place(x=450 , y=535)



# ==============================      COUNT         =================================
sb = []
font1 = ('Times', 12 , 'normal')

sv1 = IntVar()
sv2 = IntVar()
sv3 = IntVar()
sv4 = IntVar()
sv5 = IntVar()
sv6 = IntVar()
sv7 = IntVar()
sv8 = IntVar()
sv9 = IntVar()
sv10 = IntVar()
sv11 = IntVar()
sv12 = IntVar()
sv13 = IntVar()
sv14 = IntVar()
sv15 = IntVar()
sv16 = IntVar()
sv17 = IntVar()

sb1 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv1,bg='#41BC66')
sb1.place(x=30,y=155)
sb.append(sb1)

sb2 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv2,bg='#41BC66')
sb2.place(x=160,y=155)
sb.append(sb2)
sb3 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv3,bg='#41BC66')
sb3.place(x=300,y=155)
sb.append(sb3)
sb4 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv4,bg='#41BC66')
sb4.place(x=450,y=155)
sb.append(sb4)
sb5 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv5,bg='#41BC66')
sb5.place(x=30,y=330)
sb.append(sb5)
sb6 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv6,bg='#41BC66')
sb6.place(x=160,y=330)
sb.append(sb6)
sb7 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv7,bg='#41BC66')
sb7.place(x=300,y=330)
sb.append(sb7)
sb8 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv8,bg='#41BC66')
sb8.place(x=450,y=330)
sb.append(sb8)
sb9 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv9,bg='#41BC66')
sb9.place(x=30,y=488)
sb.append(sb9)
sb10 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv10,bg='#41BC66')
sb10.place(x=160,y=488)
sb.append(sb10)
sb11 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv11,bg='#41BC66')
sb11.place(x=300,y=488)
sb.append(sb11)
sb12 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv12,bg='#41BC66')
sb12.place(x=450,y=488)
sb.append(sb12)
sb13 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv13,bg='#41BC66')
sb13.place(x=450,y=488)
sb.append(sb13)
sb14 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv14,bg='#41BC66')
sb14.place(x=30,y=643)
sb.append(sb14)
sb15 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv15,bg='#41BC66')
sb15.place(x=160,y=643)
sb.append(sb15)
sb16 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv16,bg='#41BC66')
sb16.place(x=300,y=643)
sb.append(sb16)
sb17 = Spinbox(F1, from_=0, to_=5,font=font1,width=10,textvariable=sv17,bg='#41BC66')
sb17.place(x=450,y=643)
sb.append(sb17)

#=================== BUTTONS =============
b1 = Button(text='شراء ',fg='white',font=('Tajawal 15'),width=15,height=2,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',command=bil)
b1.place(x=10 ,y=725)

b2 = Button(text='فاتورة جديدة',fg='white',font=('Tajawal 15'),width=15,height=2,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',command=clea)
b2.place(x=210 ,y=725)

b3 = Button(text='إغلاق البرنامج ',fg='white',font=('Tajawal 15'),width=15,height=2,bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2')
b3.place(x=410 ,y=725)

# ============ freame 2 =======
F2 = Frame(pro , bg='gray', width=500 , height=715)
F2.place(x=605,y=1)

trv = ttk.Treeview(F2,selectmode='browse')
trv.place(x=1,y=1, width=490,height=715)
trv["columns"]=('1','2','3')
trv.column("#0",width=100,anchor='c')
trv.column("1",width=60,anchor='c')
trv.column("2",width=60,anchor='c')
trv.column("3",width=65,anchor='c')

trv.heading("#0", text=" المواد", anchor='c')
trv.heading("1", text=" السعر", anchor='c')
trv.heading("2", text=" العدد", anchor='c')
trv.heading("3", text=" الحساب الكلي", anchor='c')
#====================                 ============================
\





#==== price ===== 

menu = {
    0:['ADIDAS',20],
    1:['NIKE',35],
    2:['PUMA',35],
    3:['VL',154],
    4:['MODA',165],
    5:['FASHION',147],
    6:['SNEAKERS',85],
    7:['BOOT',180],
    8:['SNAKY',147],
    9:['SWEET',147],
    10:['NOOOR',355],
    11:['LONG B',85],
    12:['CIRIE',354],
    13:['NDI',748],
    14:['NIDIE',666],
    15:['NODUVL',100],
    16:['NBUY',148],
}





pro.mainloop()